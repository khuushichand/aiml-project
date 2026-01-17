import base64
import json
import shutil
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, timezone

import pytest
from fastapi.testclient import TestClient

from tldw_Server_API.app.core.AuthNZ.User_DB_Handling import User, get_request_user
from tldw_Server_API.app.core.config import settings
from tldw_Server_API.app.core.DB_Management.Collections_DB import CollectionsDatabase
from tldw_Server_API.app.core.DB_Management.db_path_utils import DatabasePaths


pytestmark = pytest.mark.integration

BASE_OPTIONS = {"persist": True}

@pytest.fixture()
def client_with_user(monkeypatch):
    async def override_user():
        return User(id=321, username="tester", email=None, is_active=True)

    monkeypatch.setenv("MINIMAL_TEST_APP", "0")
    monkeypatch.setenv("ULTRA_MINIMAL_APP", "0")

    base_dir = Path.cwd() / "Databases" / "test_user_dbs_files"
    shutil.rmtree(base_dir, ignore_errors=True)
    base_dir.mkdir(parents=True, exist_ok=True)
    prev_base_dir = settings.get("USER_DB_BASE_DIR")
    settings.USER_DB_BASE_DIR = str(base_dir)
    monkeypatch.setenv("USER_DB_BASE_DIR", str(base_dir))

    app = None
    try:
        from importlib import import_module, reload

        mod = import_module("tldw_Server_API.app.main")
        mod = reload(mod)
        app = mod.app
        app.dependency_overrides[get_request_user] = override_user
        with TestClient(app) as client:
            yield client
    finally:
        if app is not None:
            app.dependency_overrides.clear()
        if prev_base_dir is not None:
            settings.USER_DB_BASE_DIR = prev_base_dir
        else:
            try:
                del settings.USER_DB_BASE_DIR
            except AttributeError:
                pass


def test_create_and_export_markdown_table(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "md", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    data = response.json()
    artifact = data["artifact"]
    assert artifact["file_type"] == "markdown_table"
    assert artifact["export"]["status"] == "ready"
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "| Name | Score |" in download.text

    download_again = client_with_user.get(export_url)
    assert download_again.status_code == 409, download_again.text


def test_get_file_artifact(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact_id = response.json()["artifact"]["file_id"]

    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 200, fetch.text
    artifact = fetch.json()["artifact"]
    assert artifact["file_id"] == artifact_id
    assert artifact["export"]["status"] == "none"


def test_export_not_ready_returns_404(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact_id = response.json()["artifact"]["file_id"]

    export = client_with_user.get(f"/api/v1/files/{artifact_id}/export?format=md")
    assert export.status_code == 404, export.text


def test_async_export_returns_pending(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "md", "mode": "url", "async_mode": "async"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 202, response.text
    artifact = response.json()["artifact"]
    assert artifact["export"]["status"] == "pending"
    assert artifact["export"]["job_id"]


def test_delete_file_artifact_soft(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact_id = response.json()["artifact"]["file_id"]

    delete = client_with_user.delete(f"/api/v1/files/{artifact_id}")
    assert delete.status_code == 200, delete.text
    assert delete.json()["success"] is True

    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 404, fetch.text


def test_create_and_export_xlsx(client_with_user):
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    payload = {
        "file_type": "xlsx",
        "title": "Roster",
        "payload": {
            "sheets": [
                {"name": "Sheet1", "columns": ["Name", "Score"], "rows": [["Ada", 95]]},
            ]
        },
        "export": {"format": "xlsx", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(download.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=1, column=2).value == "Score"
    assert ws.cell(row=2, column=1).value == "Ada"
    assert ws.cell(row=2, column=2).value == 95


def test_create_and_export_csv_table(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "Name,Score" in download.text
    assert "Ada,95" in download.text


def test_create_and_export_json_table(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "json", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code in (200, 201), response.text
    data = response.json()
    artifact = data["artifact"]
    assert artifact["export"]["status"] == "ready"
    assert artifact["export"]["format"] == "json"
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    exported = download.json()
    expected_rows = [
        dict(zip(payload["payload"]["columns"], payload["payload"]["rows"][0]))
    ]
    assert exported == expected_rows


def test_create_requires_options(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_create_rejects_persist_false(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": {"persist": False},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_auto_async_uses_size_estimate(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Big",
        "payload": {"columns": ["Name", "Blob"], "rows": [["Ada", "x" * 5000]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "auto"},
        "options": {"persist": True, "max_bytes": 100},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 202, response.text
    artifact = response.json()["artifact"]
    assert artifact["export"]["status"] == "pending"


def test_create_warnings_on_duplicate_columns(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Dupes",
        "payload": {"columns": ["Name", "Name"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    warnings = response.json()["artifact"]["validation"]["warnings"]
    assert warnings
    assert warnings[0]["code"] == "duplicate_columns"


def test_inline_export_returns_content_b64(client_with_user, monkeypatch):
    monkeypatch.setenv("FILES_INLINE_MAX_BYTES", "1024")
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "md", "mode": "inline", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    export_info = artifact["export"]
    assert export_info["content_b64"]
    assert export_info["status"] == "none"
    assert export_info["url"] is None
    decoded = base64.b64decode(export_info["content_b64"]).decode("utf-8")
    assert "| Name |" in decoded
    artifact_id = artifact["file_id"]
    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 200, fetch.text
    assert fetch.json()["artifact"]["export"]["status"] == "none"
    export = client_with_user.get(f"/api/v1/files/{artifact_id}/export?format=md")
    assert export.status_code == 409, export.text


def test_inline_export_falls_back_to_url(client_with_user, monkeypatch):
    monkeypatch.setenv("FILES_INLINE_MAX_BYTES", "10")
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada Lovelace"]]},
        "export": {"format": "md", "mode": "inline", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_info = response.json()["artifact"]["export"]
    assert export_info["url"]
    assert export_info["content_b64"] is None


def test_export_expired_clears_state(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "md", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    artifact_id = artifact["file_id"]
    export_url = artifact["export"]["url"]
    assert export_url

    cdb = CollectionsDatabase.for_user(user_id=321)
    row = cdb.get_file_artifact(artifact_id)
    past = (datetime.now(timezone.utc) - timedelta(seconds=1)).replace(microsecond=0).isoformat()
    cdb.update_file_artifact_export(
        artifact_id,
        export_status="ready",
        export_format=row.export_format,
        export_storage_path=row.export_storage_path,
        export_bytes=row.export_bytes,
        export_content_type=row.export_content_type,
        export_job_id=row.export_job_id,
        export_expires_at=past,
        export_consumed_at=None,
    )

    outputs_dir = DatabasePaths.get_user_temp_outputs_dir(321)
    if row.export_storage_path:
        export_path = outputs_dir / row.export_storage_path
        assert export_path.exists()
    expired = client_with_user.get(export_url)
    assert expired.status_code == 404, expired.text
    assert expired.json().get("detail") == "export_expired"
    if row.export_storage_path:
        assert not export_path.exists()

    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 200, fetch.text
    assert fetch.json()["artifact"]["export"]["status"] == "none"


def test_invalid_export_format_returns_422(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_export_size_exceeded_returns_422(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "sync"},
        "options": {"persist": True, "max_bytes": 1},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_create_and_export_html_table(client_with_user):
    payload = {
        "file_type": "html_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "html", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_url = response.json()["artifact"]["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "<table>" in download.text
    assert "<th>Name</th>" in download.text


def test_create_and_export_ical(client_with_user):
    pytest.importorskip("icalendar", reason="icalendar not installed")
    payload = {
        "file_type": "ical",
        "title": "Schedule",
        "payload": {
            "calendar": {
                "prodid": "-//tldw//files//EN",
                "version": "2.0",
                "timezone": "UTC",
                "events": [
                    {
                        "uid": "event-1",
                        "summary": "Kickoff",
                        "start": "2026-01-01T10:00:00",
                        "end": "2026-01-01T11:00:00",
                    }
                ],
            }
        },
        "export": {"format": "ics", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_url = response.json()["artifact"]["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "BEGIN:VCALENDAR" in download.text
    assert "BEGIN:VEVENT" in download.text
